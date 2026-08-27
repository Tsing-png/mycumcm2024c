"""独立验证：清洗文件 vs 原始 xlsx 逐条比对（与清洗脚本不同代码路径）。临时脚本，不入产物。"""
import csv
import re
import openpyxl
from pathlib import Path

ROOT = Path(".")
RAW = ROOT / "workspace/data_raw"
CLEAN = ROOT / "workspace/data_clean"
errs = []


def read_csv(name):
    with open(CLEAN / name, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def strip(v):
    return None if v is None else str(v).strip()


# ---------------- 1. plots vs 附件1 地块
wb1 = openpyxl.load_workbook(RAW / "附件1.xlsx", data_only=True)
ws = wb1["乡村的现有耕地"]
raw_plots = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] is None or not isinstance(r[2], (int, float)):
        continue
    raw_plots.append((strip(r[0]), strip(r[1]), float(r[2])))
clean_plots = [(p["plot_id"], p["plot_type"], float(p["area_mu"])) for p in read_csv("plots.csv")]
if raw_plots != clean_plots:
    errs.append(f"plots 不一致: raw={len(raw_plots)} clean={len(clean_plots)}")
    for a, b in zip(raw_plots, clean_plots):
        if a != b:
            errs.append(f"  plot diff: raw={a} clean={b}")

# ---------------- 2. crops vs 附件1 作物
ws = wb1["乡村种植的农作物"]
raw_crops = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not isinstance(r[0], (int, float)):
        continue
    raw_crops.append((int(r[0]), strip(r[1]), strip(r[2])))
# 合并区锚点重算（独立于清洗脚本）
merged_anchors = {}
for rng in ws.merged_cells.ranges:
    if rng.min_col == 4 and rng.min_row == rng.max_row:
        pass
    if rng.min_col == rng.max_col:
        v = ws.cell(rng.min_row, rng.min_col).value
        if v is not None:
            merged_anchors[(rng.min_col, rng.min_row, rng.max_row)] = v
expected_suit = {}
for i in range(2, 43):
    exp = None
    for (col, r0, r1), v in merged_anchors.items():
        if col == 4 and r0 <= i <= r1:
            exp = v
    if exp is None:
        exp = ws.cell(i, 4).value
    if exp is not None:
        segs = [re.sub(r"\s+", " ", s).strip() for s in re.split(r"\n+", str(exp))]
        exp = "；".join(x for x in segs if x)
    expected_suit[i - 1] = exp
clean_crops = read_csv("crops.csv")
for c in clean_crops:
    cid = int(c["crop_id"])
    raw = raw_crops[cid - 1]
    if (cid, c["crop_name"], c["crop_category"]) != raw:
        errs.append(f"crops 不一致: clean={c} raw={raw}")
    if c["suitability"] != expected_suit[cid]:
        errs.append(f"crop{cid} suitability: clean={c['suitability']!r} expected={expected_suit[cid]!r}")
if len(clean_crops) != 41:
    errs.append(f"crops 行数={len(clean_crops)}")

# ---------------- 3. stats vs 附件2 统计
wb2 = openpyxl.load_workbook(RAW / "附件2.xlsx", data_only=True)
ws = wb2["2023年统计的相关数据"]
raw_stats = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not isinstance(r[0], (int, float)):
        continue
    m = re.fullmatch(r"\s*([\d.]+)\s*-\s*([\d.]+)\s*", str(r[7]))
    raw_stats.append((int(r[0]), int(r[1]), strip(r[2]), strip(r[3]), strip(r[4]),
                      float(r[5]), float(r[6]), strip(r[7]), float(m.group(1)), float(m.group(2))))
clean_stats = read_csv("stats_2023.csv")
if len(raw_stats) != len(clean_stats):
    errs.append(f"stats 行数 raw={len(raw_stats)} clean={len(clean_stats)}")
for a, c in zip(raw_stats, clean_stats):
    ct = (int(c["seq"]), int(c["crop_id"]), c["crop_name"], c["plot_type"], c["season"],
          float(c["yield_jin_per_mu"]), float(c["cost_yuan_per_mu"]), c["price_raw"],
          float(c["price_low"]), float(c["price_high"]))
    if a != ct:
        errs.append(f"stats 不一致: raw={a} clean={ct}")

# ---------------- 4. planting vs 附件2 种植（合并单元格独立重算）
ws = wb2["2023年的农作物种植情况"]
fill = {}
for rng in ws.merged_cells.ranges:
    if rng.min_col == 1 and rng.min_row > 1:
        fill[(rng.min_row, rng.max_row)] = ws.cell(rng.min_row, 1).value
raw_plant = []
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if r[1] is None or not isinstance(r[1], (int, float)):
        continue
    plot = r[0]
    if plot is None:
        plot = next(v for (r0, r1), v in fill.items() if r0 <= i <= r1)
    raw_plant.append((strip(plot), int(r[1]), strip(r[2]), strip(r[3]),
                      float(r[4]), strip(r[5])))
clean_plant = read_csv("planting_2023.csv")
if len(raw_plant) != len(clean_plant):
    errs.append(f"planting 行数 raw={len(raw_plant)} clean={len(clean_plant)}")
for a, c in zip(raw_plant, clean_plant):
    ct = (c["plot_id"], int(c["crop_id"]), c["crop_name"], c["crop_category"],
          float(c["area_mu"]), c["season"])
    if a != ct:
        errs.append(f"planting 不一致: raw={a} clean={ct}")

# ---------------- 5. derived = 普通大棚第一季复制
clean_der = read_csv("stats_2023_derived.csv")
src = {(int(s["crop_id"])): s for s in clean_stats if s["plot_type"] == "普通大棚" and s["season"] == "第一季"}
for d in clean_der:
    s = src[int(d["crop_id"])]
    ok = (d["plot_type"] == "智慧大棚" and d["season"] == "第一季"
          and float(d["yield_jin_per_mu"]) == float(s["yield_jin_per_mu"])
          and float(d["cost_yuan_per_mu"]) == float(s["cost_yuan_per_mu"])
          and float(d["price_low"]) == float(s["price_low"])
          and float(d["price_high"]) == float(s["price_high"]))
    if not ok:
        errs.append(f"derived 不一致: crop{d['crop_id']} {d} vs src {s}")
if len(clean_der) != 18:
    errs.append(f"derived 行数={len(clean_der)}")

# ---------------- 6. 模板结构 vs 附件3（抽查 result1_2/result2 与 result1_1 一致 + 行序）
import json
tmpl = json.load(open(CLEAN / "template_structure.json", encoding="utf-8"))
wb3 = openpyxl.load_workbook(RAW / "附件3/result2.xlsx", data_only=True)
ws3 = wb3["2030"]
raw_rows = list(ws3.iter_rows(values_only=True))
raw_first = [strip(r[1]) for r in raw_rows[1:55]]
raw_second = [strip(r[1]) for r in raw_rows[55:83]]
if tmpl["result2.xlsx"]["2030"]["first_season_plots"] != raw_first:
    errs.append("模板 2030 第一季行序不一致")
if tmpl["result2.xlsx"]["2030"]["second_season_plots"] != raw_second:
    errs.append("模板 2030 第二季行序不一致")
hdr = [strip(c) for c in raw_rows[0][2:43]]
crop_names = [c["crop_name"] for c in read_csv("crops.csv")]
if hdr != crop_names:
    errs.append("模板表头作物序与 crops.csv 不一致")

print("errors:", len(errs))
for e in errs[:20]:
    print(" -", e)
print("VERIFY_OK" if not errs else "VERIFY_FAILED")
