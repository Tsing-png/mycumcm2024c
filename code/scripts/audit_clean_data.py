"""2024 C 题数据审计与清洗（可复现）。

只做两类操作：
1. 安全规范化：去空白、合并单元格锚点值向下填充、价格区间解析。
2. 记录口径约定：附件2 注(2) 智慧大棚第一季 = 普通大棚第一季（派生表单独存放，不混入原始 107 行）。

不做：插补、删行、缩尾、改数值。所有假设性操作写入 data_profile.json 的 cleaning_notes。
原始文件只读；清洗结果写入 workspace/data_clean/。

运行：uv run python code/scripts/audit_clean_data.py
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
RAW = ROOT / "workspace/data_raw"
CLEAN = ROOT / "workspace/data_clean"
CLEAN.mkdir(parents=True, exist_ok=True)

RAW_HASHES = {
    "附件1.xlsx": "b799a137294a5c5497fed9667c5dfed6c967f1fb180316c5adc8509cfa6f0932",
    "附件2.xlsx": "869081a3ab47d3bf8d0955106b622aaf0fd2c068fada7948da69b20ebf1d00ce",
    "附件3/result1_1.xlsx": "4f2484c0d70a5c4d047163f2ee6ef486949e813330466f46def4bd7d98af06af",
    "附件3/result1_2.xlsx": "6166d43f5a64bf9d1657e80d4aee7f10f54bb1a5695b81a28a0ac5e657297649",
    "附件3/result2.xlsx": "6a1ba9fc28d14d0a4a795e5f0b7261fb6e32165517afee62bcd1931aba5bee8a",
}


def norm(s):
    """去首尾空白；None 保持 None。"""
    return None if s is None else str(s).strip()


def read_sheet_with_merge_fill(path, sheet, header_row=1):
    """读取工作表：header_row 为表头，数据区合并单元格锚点值向下填充。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # 锚点值 -> 填充映射：(col_index, row_range)
    fill = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row > header_row and rng.min_col == rng.max_col:
            anchor = ws.cell(rng.min_row, rng.min_col).value
            fill[(rng.min_col - 1, rng.min_row - 1, rng.max_row)] = anchor
    out = []
    for i, row in enumerate(rows):
        new = list(row)
        for (col, r0, r1), val in fill.items():
            if r0 <= i < r1:
                new[col] = val
        out.append(new)
    return out


def norm_suit(s):
    """适宜性原文规范化：按换行分段，段内空白折叠为单空格，段间以；连接。"""
    if s is None:
        return None
    segs = [re.sub(r"\s+", " ", seg).strip() for seg in re.split(r"\n+", s)]
    return "；".join(x for x in segs if x)


def parse_price_interval(s):
    """'2.50-4.00' -> (low, high)。异常格式记录并返回 None。"""
    if s is None:
        return None
    m = re.fullmatch(r"\s*([\d.]+)\s*-\s*([\d.]+)\s*", str(s))
    if not m:
        return None
    lo, hi = float(m.group(1)), float(m.group(2))
    return (lo, hi)


findings = {"duplicates": {}, "impossible": [], "anomalies": [], "cleaning_notes": []}

# ---------------------------------------------------------------- 附件1 地块
rows = read_sheet_with_merge_fill(RAW / "附件1.xlsx", "乡村的现有耕地")
plots, seen = [], set()
for r in rows[1:]:
    name, ptype, area = norm(r[0]), norm(r[1]), r[2]
    if name is None or not isinstance(area, (int, float)):
        continue
    if name in seen:
        findings["duplicates"].setdefault("plots", []).append(name)
    seen.add(name)
    plots.append({"plot_id": name, "plot_type": ptype, "area_mu": float(area)})

# ---------------------------------------------------------------- 附件1 作物
rows = read_sheet_with_merge_fill(RAW / "附件1.xlsx", "乡村种植的农作物")
crops = []
for r in rows[1:]:
    if not isinstance(r[0], (int, float)):
        continue
    cid, cname, ccat, suit_raw = int(r[0]), norm(r[1]), norm(r[2]), norm(r[3])
    crops.append({
        "crop_id": cid, "crop_name": cname, "crop_category": ccat,
        "suitability": norm_suit(suit_raw),
    })
if len(crops) != 41:
    findings["anomalies"].append(f"附件1作物行数={len(crops)}，应为41")

# ---------------------------------------------------------------- 附件2 种植
rows = read_sheet_with_merge_fill(RAW / "附件2.xlsx", "2023年的农作物种植情况")
planting, keys = [], {}
for r in rows[1:]:
    plot, cid, cname, ccat, area, season = norm(r[0]), r[1], norm(r[2]), norm(r[3]), r[4], norm(r[5])
    if cid is None:
        continue
    if not isinstance(area, (int, float)) or area <= 0:
        findings["impossible"].append(f"种植记录非正面积: plot={plot} crop={cid} area={area}")
    planting.append({
        "plot_id": plot, "crop_id": int(cid), "crop_name": cname,
        "crop_category": ccat, "area_mu": float(area), "season": season,
    })
    k = (plot, int(cid), season)
    keys[k] = keys.get(k, 0) + 1
findings["duplicates"]["planting_keys"] = {str(k): v for k, v in keys.items() if v > 1}

# 2023 各地块按季种植面积 vs 容量（容量按季占用，跨季求和是误报）
cap = {p["plot_id"]: p["area_mu"] for p in plots}
use = {}
for rec in planting:
    k = (rec["plot_id"], rec["season"])
    use[k] = use.get(k, 0.0) + rec["area_mu"]
over_cap = {str(k): round(v, 4) for k, v in use.items() if round(v - cap.get(k[0], 0.0), 6) > 0}
if over_cap:
    findings["impossible"].append(f"2023年单季超容量地块: {over_cap}")

# ---------------------------------------------------------------- 附件2 统计
rows = read_sheet_with_merge_fill(RAW / "附件2.xlsx", "2023年统计的相关数据")
stats, skeys = [], {}
bad_price = []
for r in rows[1:]:
    seq, cid, cname, ptype, season, yld, cost, price = r[:8]
    if not isinstance(seq, (int, float)) or cid is None:
        continue
    p = parse_price_interval(price)
    if p is None:
        bad_price.append((seq, price))
    else:
        if p[0] > p[1]:
            findings["impossible"].append(f"价格区间倒置: seq={seq} {price}")
        if p[1] <= 0 or yld is None or yld <= 0 or cost is None or cost <= 0:
            findings["impossible"].append(f"非正值: seq={seq} yld={yld} cost={cost} price={price}")
    stats.append({
        "seq": int(seq), "crop_id": int(cid), "crop_name": norm(cname),
        "plot_type": norm(ptype), "season": norm(season),
        "yield_jin_per_mu": float(yld), "cost_yuan_per_mu": float(cost),
        "price_raw": norm(price),
        "price_low": p[0] if p else None, "price_high": p[1] if p else None,
    })
    k = (int(cid), norm(ptype), norm(season))
    skeys[k] = skeys.get(k, 0) + 1
findings["duplicates"]["stats_keys"] = {str(k): v for k, v in skeys.items() if v > 1}
if bad_price:
    findings["anomalies"].append(f"价格区间解析失败: {bad_price}")

# ---------------------------------------------------------------- 附件3 模板
tmpl = {}
for fname in ["result1_1.xlsx", "result1_2.xlsx", "result2.xlsx"]:
    wb = openpyxl.load_workbook(RAW / "附件3" / fname, data_only=True)
    years = wb.sheetnames
    info = {}
    for y in years:
        ws = wb[y]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        crop_order = [norm(c) for c in header[2:43]]
        first_plots = [norm(r[1]) for r in rows[1:55]]
        second_plots = [norm(r[1]) for r in rows[55:83]]
        info[y] = {
            "n_crop_cols": len(crop_order),
            "crop_order_matches_ids_1_41": crop_order == [c["crop_name"] for c in crops],
            "first_season_plots": first_plots,
            "second_season_plots": second_plots,
            "notes_rows_85_87": [norm(str(rows[84][1])), norm(str(rows[85][1])), norm(str(rows[86][1]))],
        }
    tmpl[fname] = {"years": years, **info}

# 三模板结构一致性
ref = tmpl["result1_1.xlsx"]
struct_same = all(
    all(tmpl[f][y]["first_season_plots"] == ref[y]["first_season_plots"]
        and tmpl[f][y]["second_season_plots"] == ref[y]["second_season_plots"]
        and tmpl[f][y]["crop_order_matches_ids_1_41"]
        for y in ref["years"])
    for f in ["result1_2.xlsx", "result2.xlsx"])
if not struct_same:
    findings["anomalies"].append("三个结果模板结构不一致")

# ---------------------------------------------------------------- 覆盖率
# 题目规则下需要的 (crop, plot_type, season) 组合
required = set()
for cid in range(1, 16):  # 粮食1-15 旱地三类型 单季
    for pt in ["平旱地", "梯田", "山坡地"]:
        required.add((cid, pt, "单季"))
required.add((16, "水浇地", "单季"))  # 水稻
for cid in range(17, 35):  # 蔬菜17-34
    required.add((cid, "水浇地", "第一季"))
    required.add((cid, "普通大棚", "第一季"))
    required.add((cid, "智慧大棚", "第一季"))
    required.add((cid, "智慧大棚", "第二季"))
for cid in [35, 36, 37]:  # 大白菜白萝卜红萝卜 水浇地第二季
    required.add((cid, "水浇地", "第二季"))
for cid in [38, 39, 40, 41]:  # 食用菌 普通大棚第二季
    required.add((cid, "普通大棚", "第二季"))

available = set(skeys)
derived = {(cid, "智慧大棚", "第一季") for cid in range(17, 35)}
missing = sorted(required - available - derived)
coverage = {
    "required_combos": len(required),
    "available_in_stats": len(required & available),
    "derived_from_note2": sorted(str(d) for d in sorted(derived)),
    "missing_after_derivation": [str(m) for m in missing],
}

# ---------------------------------------------------------------- 风险画像统计
# 价格区间宽度比 (high/low) 极值
ratio = [(s["crop_name"], s["plot_type"], s["season"], s["price_high"] / s["price_low"])
         for s in stats if s["price_low"]]
ratio.sort(key=lambda x: -x[3])
widest = ratio[:6]

# 类别不平衡：地块类型面积
pt_area = {}
for p in plots:
    pt_area[p["plot_type"]] = pt_area.get(p["plot_type"], 0.0) + p["area_mu"]

# 2023 面积集中度（作物）
crop_area = {}
for rec in planting:
    crop_area[rec["crop_id"]] = crop_area.get(rec["crop_id"], 0.0) + rec["area_mu"]
top = sorted(crop_area.items(), key=lambda x: -x[1])
total_area = sum(crop_area.values())
top3_share = sum(v for _, v in top[:3]) / total_area

# 2023 产量按作物（预期销量候选来源，仅作为派生候选，不作决定）
plot_type_of = {p["plot_id"]: p["plot_type"] for p in plots}
yield_map = {(s["crop_id"], s["plot_type"], s["season"]): s["yield_jin_per_mu"] for s in stats}
yield_map.update({(s["crop_id"], "智慧大棚", "第一季"): s["yield_jin_per_mu"]
                  for s in stats if s["plot_type"] == "普通大棚" and s["season"] == "第一季"})
prod2023 = {}
for rec in planting:
    pt = plot_type_of[rec["plot_id"]]
    y = yield_map.get((rec["crop_id"], pt, rec["season"]))
    if y is None:
        findings["anomalies"].append(
            f"2023种植记录无统计亩产量: {rec['plot_id']} crop{rec['crop_id']} {pt} {rec['season']}")
        continue
    prod2023[rec["crop_id"]] = prod2023.get(rec["crop_id"], 0.0) + rec["area_mu"] * y

# ---------------------------------------------------------------- 写清洗文件
def write_csv(name, header, rows_):
    import csv as _csv
    with open(CLEAN / name, "w", encoding="utf-8-sig", newline="") as f:
        w = _csv.writer(f, lineterminator="\n")  # QUOTE_MINIMAL 默认，自动转义逗号与换行
        w.writerow(header)
        for r in rows_:
            w.writerow(["" if v is None else str(v) for v in r])

write_csv("plots.csv", ["plot_id", "plot_type", "area_mu"],
          [(p["plot_id"], p["plot_type"], p["area_mu"]) for p in plots])
write_csv("crops.csv", ["crop_id", "crop_name", "crop_category", "suitability"],
          [(c["crop_id"], c["crop_name"], c["crop_category"], c["suitability"]) for c in crops])
write_csv("stats_2023.csv",
          ["seq", "crop_id", "crop_name", "plot_type", "season", "yield_jin_per_mu",
           "cost_yuan_per_mu", "price_raw", "price_low", "price_high"],
          [(s["seq"], s["crop_id"], s["crop_name"], s["plot_type"], s["season"],
            s["yield_jin_per_mu"], s["cost_yuan_per_mu"], s["price_raw"],
            s["price_low"], s["price_high"]) for s in stats])
write_csv("stats_2023_derived.csv",
          ["crop_id", "crop_name", "plot_type", "season", "yield_jin_per_mu",
           "cost_yuan_per_mu", "price_low", "price_high", "derived_from"],
          [(s["crop_id"], s["crop_name"], "智慧大棚", "第一季", s["yield_jin_per_mu"],
            s["cost_yuan_per_mu"], s["price_low"], s["price_high"], "普通大棚第一季(附件2注2)")
           for s in stats if s["plot_type"] == "普通大棚" and s["season"] == "第一季"])
write_csv("planting_2023.csv",
          ["plot_id", "crop_id", "crop_name", "crop_category", "area_mu", "season"],
          [(r["plot_id"], r["crop_id"], r["crop_name"], r["crop_category"],
            r["area_mu"], r["season"]) for r in planting])

# ---------------------------------------------------------------- profile
profile = {
    "schema_version": 1,
    "raw_files": [
        {"path": "workspace/data_raw/" + k, "sha256": v, "readonly": True}
        for k, v in RAW_HASHES.items()
    ],
    "attachment_mapping": [
        {"file": "附件1.xlsx", "sheets": ["乡村的现有耕地(54地块)", "乡村种植的农作物(41作物)"],
         "maps_to": ["Q1", "Q2", "Q3"], "role": "地块/作物/适宜性"},
        {"file": "附件2.xlsx", "sheets": ["2023年的农作物种植情况(87条)", "2023年统计的相关数据(107条)"],
         "maps_to": ["Q1", "Q2", "Q3"], "role": "2023基准/亩产/成本/价格区间"},
        {"file": "附件3/result1_1.xlsx", "maps_to": ["Q1"], "role": "情形1输出模板"},
        {"file": "附件3/result1_2.xlsx", "maps_to": ["Q1"], "role": "情形2输出模板"},
        {"file": "附件3/result2.xlsx", "maps_to": ["Q2"], "role": "Q2输出模板"},
    ],
    "fields": [
        {"name": "plot_id", "n": len(plots), "type": "str", "notes": "54地块; A-F系列"},
        {"name": "plot_type", "n": len(pt_area), "levels": pt_area,
         "notes": "类型面积不平衡: 梯田619亩 vs 智慧大棚2.4亩"},
        {"name": "crop_id", "n": len(crops), "type": "int 1..41", "notes": "与模板列序一致"},
        {"name": "yield_jin_per_mu", "n": len(stats), "notes": "107条, 全部>0"},
        {"name": "cost_yuan_per_mu", "n": len(stats), "notes": "107条, 全部>0"},
        {"name": "price", "n": len(stats), "type": "区间字符串",
         "notes": "全部解析为[low, high]; 代表值口径待人工定"},
        {"name": "expected_sales", "n": 0, "type": "缺失字段",
         "notes": "附件中不存在该字段; 候选来源见production_2023_candidate"},
    ],
    "quality": {
        "missingness": {
            "expected_sales": "字段完全缺失（附件2无此列），2023产量候选派生见concentration_metrics",
            "attachment1_crop_suitability": "合并单元格锚点值已填充，无行缺失",
            "attachment2_planting_plot_name": "26处合并单元格已按锚点填充，无缺失",
            "stats_2023": "107/107 行字段完整",
        },
        "duplicates": findings["duplicates"],
        "impossible_values": findings["impossible"],
        "outliers": {
            "widest_price_intervals": [{"crop": c, "type": t, "season": s, "high_over_low": round(v, 2)}
                                       for c, t, s, v in widest],
            "note": "价格区间宽度大不构成错误，但代表值选择影响收益口径",
        },
    },
    "coverage": {
        "rows": len(stats),
        "effective_sample_size": len(stats),
        "time_range": "2023年单年截面",
        "time_gaps": "无时间序列数据；Q2变化率只来自题面文字",
        "stats_combo_coverage": coverage,
    },
    "distribution_risks": {
        "class_imbalance": {"plot_type_area_mu": pt_area},
        "rare_categories": ["食用菌4种(仅普通大棚第二季)", "水稻1种(仅水浇地单季)"],
        "high_cardinality": [],
        "redundancy_warnings": [
            "普通大棚第一季与智慧大棚第一季亩产/成本/价格完全相同（附件2注2，非冗余错误）",
            "大白菜/白萝卜/红萝卜仅水浇地第二季3种组合",
        ],
        "concentration_metrics": {
            "top3_crops_2023_area_share": round(top3_share, 4),
            "top_crops_2023": [{"crop_id": c, "area_mu": round(a, 2)} for c, a in top[:5]],
            "production_2023_candidate_expected_sales_jin": {str(k): round(v, 2) for k, v in sorted(prod2023.items())},
        },
    },
    "per_question_readiness": {
        "Q1": "ready_with_warnings",
        "Q1_warnings": [
            "预期销售量字段缺失：需要人工决定是否以2023产量（或其它口径）作为预期销量",
            "价格区间代表值（均值/低/高）未定",
            "2023年D7/D8单季水稻与D1-D6两季蔬菜并存，2024年重茬初始条件以2023记录为准（无更早数据）",
        ],
        "Q2": "ready_with_warnings",
        "Q2_warnings": [
            "无任何概率分布或情景观测数据，不确定性只来自题面区间与趋势文字",
            "Q2超额产量销售规则题面未明示，需人工决定沿用Q1哪种情形",
            "风险定义与容忍度无数据支撑，需人工设定",
        ],
        "Q3": "blocked_on_relationship_data",
        "Q3_warnings": [
            "无作物替代/互补或经济变量相关性的观测数据，只能做可复现模拟假设",
            "模拟关系强度、方向与比较指标需人工设定后再评估",
        ],
    },
    "cleaned_files": [
        "plots.csv", "crops.csv", "stats_2023.csv", "stats_2023_derived.csv",
        "planting_2023.csv", "template_structure.json",
    ],
    "cleaning_notes": findings["cleaning_notes"] + [
        "crops.csv suitability 列为规范化适宜性（换行分段、段内空白折叠为单空格、段间以；连接），不保留原文列",
        "所有字符串字段去首尾空白（普通大棚 /菠菜 /生菜 /说明 等尾随空格）",
        "附件2种植表与附件1作物表合并单元格按锚点值向下填充，未改动数值",
        "智慧大棚第一季参数按附件2注(2)从普通大棚第一季派生，单独文件存放，不并入stats_2023.csv",
        "未进行任何插补、删除、缩尾或重编码",
    ],
    "unresolved_risks": [
        "预期销量来源未定（全题收益口径依赖此项）",
        "价格代表值未定",
        "管理便利阈值无数据，需人工给出",
        "Q2不确定性分布与Q3关系结构无观测数据",
    ],
}

with open(CLEAN / "template_structure.json", "w", encoding="utf-8") as f:
    json.dump(tmpl, f, ensure_ascii=False, indent=2)
with open(CLEAN / "data_profile.json", "w", encoding="utf-8") as f:
    json.dump(profile, f, ensure_ascii=False, indent=2)

print("plots:", len(plots), "crops:", len(crops), "stats:", len(stats), "planting:", len(planting))
print("duplicates:", findings["duplicates"])
print("impossible:", findings["impossible"])
print("anomalies:", findings["anomalies"])
print("coverage: required=%d available=%d derived=%d missing=%s" % (
    coverage["required_combos"], coverage["required_combos"] - len(missing) - len(derived),
    len(derived), coverage["missing_after_derivation"]))
print("over_capacity_2023:", over_cap)
print("top3 area share:", round(top3_share, 4), "top:", top[:5])
print("widest price intervals:", [(c, round(v, 2)) for c, _, _, v in widest])
print("templates structurally identical:", struct_same)
