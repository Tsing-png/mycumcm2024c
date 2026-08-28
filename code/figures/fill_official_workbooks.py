from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
RAW_TEMPLATES = ROOT / "workspace/data_raw/附件3"


def fill_schedule(template: Path, schedule_csv: Path, output: Path) -> dict[str, float | int]:
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)

    schedule = pd.read_csv(schedule_csv, encoding="utf-8-sig")
    schedule = schedule[schedule["area_mu"] > 1e-8].copy()
    workbook = openpyxl.load_workbook(output)
    written = 0
    written_area = 0.0

    for year, year_rows in schedule.groupby("year"):
        worksheet = workbook[str(int(year))]
        crop_columns = {crop_id: crop_id + 2 for crop_id in range(1, 42)}
        first_rows = {
            str(worksheet.cell(row, 2).value): row
            for row in range(2, 56)
            if worksheet.cell(row, 2).value is not None
        }
        second_rows = {
            str(worksheet.cell(row, 2).value): row
            for row in range(56, 84)
            if worksheet.cell(row, 2).value is not None
        }

        for record in year_rows.itertuples(index=False):
            row_map = second_rows if record.season == "第二季" else first_rows
            if record.plot_id not in row_map:
                raise ValueError(
                    f"Template has no row for year={year}, season={record.season}, plot={record.plot_id}"
                )
            row = row_map[record.plot_id]
            column = crop_columns[int(record.crop_id)]
            existing = worksheet.cell(row, column).value
            if existing not in (None, ""):
                raise ValueError(
                    f"Refusing to overwrite non-empty cell {worksheet.title}!{worksheet.cell(row, column).coordinate}"
                )
            worksheet.cell(row, column).value = round(float(record.area_mu), 6)
            written += 1
            written_area += float(record.area_mu)

    workbook.save(output)
    return {"written_cells": written, "written_area_mu": written_area}


def verify_workbook(workbook_path: Path, schedule_csv: Path) -> dict[str, float | int]:
    schedule = pd.read_csv(schedule_csv, encoding="utf-8-sig")
    schedule = schedule[schedule["area_mu"] > 1e-8].copy()
    expected = {
        (int(row.year), row.season, row.plot_id, int(row.crop_id)): float(row.area_mu)
        for row in schedule.itertuples(index=False)
    }
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    checked = 0
    maximum_error = 0.0

    for (year, season, plot_id, crop_id), expected_area in expected.items():
        worksheet = workbook[str(year)]
        row_range = range(56, 84) if season == "第二季" else range(2, 56)
        row = next(r for r in row_range if worksheet.cell(r, 2).value == plot_id)
        actual = float(worksheet.cell(row, crop_id + 2).value)
        maximum_error = max(maximum_error, abs(actual - expected_area))
        checked += 1

    return {"checked_cells": checked, "maximum_absolute_error": maximum_error}


def fill_q1() -> None:
    destination = ROOT / "results/Q1/deliverables"
    jobs = [
        (
            "result1_1.xlsx",
            ROOT / "results/Q1/experiments/round2/tables/q1_m1_alpha0_share0_k3_schedule.csv",
        ),
        (
            "result1_2.xlsx",
            ROOT / "results/Q1/experiments/round2/tables/q1_m1_alpha05_share10_k3_schedule.csv",
        ),
    ]
    for workbook_name, schedule_csv in jobs:
        output = destination / workbook_name
        fill_metrics = fill_schedule(RAW_TEMPLATES / workbook_name, schedule_csv, output)
        verify_metrics = verify_workbook(output, schedule_csv)
        print(workbook_name, fill_metrics, verify_metrics)


def fill_q2() -> None:
    workbook_name = "result2.xlsx"
    schedule_csv = ROOT / "results/Q2/experiments/round3/tables/q2_m1_schedule.csv"
    output = ROOT / "results/Q2/deliverables" / workbook_name
    fill_metrics = fill_schedule(RAW_TEMPLATES / workbook_name, schedule_csv, output)
    verify_metrics = verify_workbook(output, schedule_csv)
    print(workbook_name, fill_metrics, verify_metrics)


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill copied official workbooks from verified schedules.")
    parser.add_argument("--question", choices=["Q1", "Q2", "all"], default="all")
    args = parser.parse_args()
    if args.question in {"Q1", "all"}:
        fill_q1()
    if args.question in {"Q2", "all"}:
        fill_q2()


if __name__ == "__main__":
    main()
