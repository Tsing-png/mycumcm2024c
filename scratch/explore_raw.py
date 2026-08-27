"""Temporary exploration of raw attachment structure. Not a deliverable."""
import openpyxl
from pathlib import Path

RAW = Path("workspace/data_raw")


def dump(path: Path, max_rows: int = 12):
    print("=" * 90)
    print("FILE:", path)
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"\n--- sheet: {ws.title}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
        print("merged:", [str(m) for m in ws.merged_cells.ranges][:15],
              f"(... total {len(ws.merged_cells.ranges)})" if len(ws.merged_cells.ranges) > 15 else "")
        rows = list(ws.iter_rows(values_only=True))
        for i, r in enumerate(rows[:max_rows]):
            print(f"  r{i+1}: {r}")
        if len(rows) > max_rows:
            print(f"  ... ({len(rows)} rows total)")
        # column letter widths sample: check for stray content beyond header zone
        nonempty_cols = {}
        for r in rows:
            for j, v in enumerate(r):
                if v is not None and str(v).strip() != "":
                    nonempty_cols[j] = nonempty_cols.get(j, 0) + 1
        print("  nonempty per col index:", dict(sorted(nonempty_cols.items())))


for f in [
    RAW / "附件1.xlsx",
    RAW / "附件2.xlsx",
    RAW / "附件3/result1_1.xlsx",
    RAW / "附件3/result1_2.xlsx",
    RAW / "附件3/result2.xlsx",
]:
    dump(f)
